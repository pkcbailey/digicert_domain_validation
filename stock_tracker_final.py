
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Example data loading — replace this with actual CSV read if needed
# df = pd.read_csv("your_input.csv")
data = [
    {"Ticker": "AAPL", "Shares": 50, "Purchase Date": "2025-04-16", "Purchase Price": 196.17},
    {"Ticker": "COST", "Shares": 25, "Purchase Date": "2025-04-10", "Purchase Price": 941.81},
    {"Ticker": "BRK-B", "Shares": 24, "Purchase Date": "2025-04-11", "Purchase Price": 515.75}
]

df = pd.DataFrame(data)
df = df.sort_values(by='Ticker')

# Export to Excel
output_path = "stock_tracker_report.xlsx"
df.to_excel(output_path, index=False)

# Apply 16pt font and auto-widths using openpyxl
wb = load_workbook(output_path)
ws = wb.active
font = Font(size=16)

for row in ws.iter_rows():
    for cell in row:
        cell.font = font

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(output_path)
