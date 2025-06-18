import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime, timedelta
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# CONFIGURATION
API_KEY = 'd09t8r9r01qus8rfa1ogd09t8r9r01qus8rfa1p0'
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587
EMAIL_USER = 'pkcbailey@gmail.com'
EMAIL_PASS = 'mrqc lkgy vrqv evmn'
TO_EMAIL = 'pkcbailey@gmail.com'
CSV_FILE = 'stocks.csv'
BASE_URL = "https://finnhub.io/api/v1"
EXCEL_FILE = "stock_summary.xlsx"

def get_stock_news(symbol, from_date, to_date):
    """Fetch news for a given stock symbol"""
    endpoint = f"{BASE_URL}/company-news"
    params = {
        "symbol": symbol,
        "from": from_date,
        "to": to_date,
        "token": API_KEY
    }
    
    try:
        response = requests.get(endpoint, params=params)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching news for {symbol}: {e}")
        return []

def get_premarket_data(symbol):
    """Fetch premarket data for a given stock symbol"""
    endpoint = f"{BASE_URL}/quote"
    params = {
        "symbol": symbol,
        "token": API_KEY
    }
    
    try:
        response = requests.get(endpoint, params=params)
        response.raise_for_status()
        data = response.json()
        
        # Calculate premarket change
        if data['pc'] > 0:  # Previous close
            premarket_change = data['c'] - data['pc']  # Current price - Previous close
            premarket_change_percent = (premarket_change / data['pc']) * 100
        else:
            premarket_change = 0
            premarket_change_percent = 0
            
        return {
            'current_price': data['c'],
            'previous_close': data['pc'],
            'premarket_change': premarket_change,
            'premarket_change_percent': premarket_change_percent,
            'high': data['h'],
            'low': data['l'],
            'volume': data['t']
        }
    except requests.exceptions.RequestException as e:
        print(f"Error fetching premarket data for {symbol}: {e}")
        return None

def create_excel_summary(tickers_data):
    """Create an Excel workbook with premarket data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Premarket Summary"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = [
        "Symbol", "Current Price", "Previous Close", "Change ($)", "Change (%)",
        "High", "Low", "Volume"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row, data in enumerate(tickers_data, 2):
        # Premarket data
        ws.cell(row=row, column=1, value=data['symbol']).border = border
        ws.cell(row=row, column=2, value=data['premarket']['current_price']).border = border
        ws.cell(row=row, column=3, value=data['premarket']['previous_close']).border = border
        ws.cell(row=row, column=4, value=data['premarket']['premarket_change']).border = border
        ws.cell(row=row, column=5, value=data['premarket']['premarket_change_percent']).border = border
        ws.cell(row=row, column=6, value=data['premarket']['high']).border = border
        ws.cell(row=row, column=7, value=data['premarket']['low']).border = border
        ws.cell(row=row, column=8, value=data['premarket']['volume']).border = border
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save the workbook
    wb.save(EXCEL_FILE)
    return EXCEL_FILE

def format_news_summary(tickers_data):
    """Format news data into a readable text summary"""
    news_summary = "\n📰 Recent News:\n\n"
    
    for data in tickers_data:
        if data['news']:
            news_summary += f"=== {data['symbol']} ===\n"
            for news in data['news'][:3]:  # Limit to 3 headlines per stock
                date = datetime.fromtimestamp(news['datetime']).strftime('%Y-%m-%d')
                news_summary += f"{news['headline']} ({date})\n{news['url']}\n\n"
    
    return news_summary

def main():
    # Get date range
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    from_date = yesterday.strftime('%Y-%m-%d')
    to_date = today.strftime('%Y-%m-%d')

    # Load tickers
    try:
        with open(CSV_FILE, newline='') as csvfile:
            tickers = [row[0].strip().upper() for row in csv.reader(csvfile)]
    except FileNotFoundError:
        print(f"Error: {CSV_FILE} not found")
        return

    # Fetch data for all tickers
    tickers_data = []
    for ticker in tickers:
        news_data = get_stock_news(ticker, from_date, to_date)
        premarket_data = get_premarket_data(ticker)
        if premarket_data:
            tickers_data.append({
                'symbol': ticker,
                'premarket': premarket_data,
                'news': news_data
            })

    # Create Excel summary
    excel_file = create_excel_summary(tickers_data)
    
    # Format news summary
    news_summary = format_news_summary(tickers_data)

    # Send email with Excel attachment and news in body
    msg = MIMEMultipart()
    msg['Subject'] = f"📈 Daily Stock Summary ({today.strftime('%Y-%m-%d')})"
    msg['From'] = EMAIL_USER
    msg['To'] = TO_EMAIL

    # Add body text with news
    body = f"""Daily Stock Summary

Please find attached the premarket data spreadsheet.

{news_summary}
"""
    msg.attach(MIMEText(body, 'plain'))

    # Add Excel attachment
    with open(excel_file, 'rb') as f:
        excel_attachment = MIMEApplication(f.read(), _subtype='xlsx')
        excel_attachment.add_header('Content-Disposition', 'attachment', filename=excel_file)
        msg.attach(excel_attachment)

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            server.sendmail(EMAIL_USER, TO_EMAIL, msg.as_string())
        print("Email with Excel summary sent successfully!")
    except Exception as e:
        print(f"Error sending email: {e}")

if __name__ == "__main__":
    main()


